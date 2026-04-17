import openpyxl, json

wb = openpyxl.load_workbook('blp work.xlsx', data_only=True)
data = {}
for sheet in wb.sheetnames:
    if sheet == 'map':
        continue
    ws = wb[sheet]
    symbols = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sym = str(row[1]).strip() if row[1] else None
        name = str(row[2]).strip() if row[2] else None
        if sym and sym != '#N/A' and sym != 'None' and len(sym) > 1:
            symbols.append({'symbol': sym, 'name': name if name and name != '#N/A' else sym})
    data[sheet] = symbols

for k, v in data.items():
    first5 = [x['symbol'] for x in v[:5]]
    print(k + ': ' + str(len(v)) + ' stocks -> ' + str(first5))

with open('sheet_data.json', 'w') as f:
    json.dump(data, f, indent=2)
print('Saved sheet_data.json')
