import openpyxl, json

wb = openpyxl.load_workbook('blp work.xlsx', data_only=True)

result = {}
for shname in wb.sheetnames:
    if shname == 'map':
        continue
    ws = wb[shname]
    stocks = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        nse  = str(r[1]).strip() if r[1] else ''
        name = str(r[2]).strip() if r[2] else ''
        if not nse or nse in ('None', '#N/A', 'NSE Symbol', ' NSE Symbol'):
            continue
        if not name or name in ('None', '#N/A', 'Name'):
            name = nse
        stocks.append({'symbol': nse, 'name': name})
    result[shname] = stocks
    syms = [s['symbol'] for s in stocks[:4]]
    print(f"{shname:12s}  {len(stocks):3d} stocks   {syms}")

with open('sheet_data.json', 'w') as f:
    json.dump(result, f, indent=2)
print(f"\nWritten sheet_data.json with {len(result)} sheets")
